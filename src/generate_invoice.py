#!/usr/bin/env python3
"""
Professional Invoice Generator

Generates professional invoices for validator rewards with:
- Company logo
- Total earnings calculation
- Rate of return analysis
- Professional Excel formatting
"""

import pandas as pd
import argparse
import requests
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing import image
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
import tempfile
import os
from reward_utils import adjust_reward, get_validator_type_label, get_bonded_principal


class InvoiceGenerator:
    def __init__(self, parquet_file, logo_url="https://liblockchain.org/libc-logo.png"):
        self.parquet_file = parquet_file
        self.logo_url = logo_url
        self.logo_path = None

        # Professional color scheme
        self.colors = {
            'header': 'FF2E75B6',  # Blue
            'accent': 'FF8DB4E2',  # Light blue
            'text': 'FF000000',    # Black
            'background': 'FFF2F2F2'  # Light gray
        }

    def download_logo(self):
        """Download company logo."""
        try:
            response = requests.get(self.logo_url)
            response.raise_for_status()

            # Create temp file for logo
            temp_dir = Path(tempfile.gettempdir())
            self.logo_path = temp_dir / "libc_logo.png"

            with open(self.logo_path, 'wb') as f:
                f.write(response.content)

            print(f"✅ Downloaded logo to {self.logo_path}")
            return True

        except Exception as e:
            print(f"⚠️  Could not download logo: {e}")
            return False

    def load_data(self, start_epoch, end_epoch):
        """Load and filter parquet data."""
        if not Path(self.parquet_file).exists():
            raise FileNotFoundError(f"Parquet file not found: {self.parquet_file}")

        df = pd.read_parquet(self.parquet_file)

        # Filter by epoch range
        df_filtered = df[(df['epoch'] >= start_epoch) & (df['epoch'] <= end_epoch)].copy()

        return df_filtered

    def calculate_earnings(self, df):
        """Calculate total earnings and breakdown, separating exits from earnings."""
        df_calc = df.copy()

        # Check if is_exit column exists
        has_exit_flag = 'is_exit' in df_calc.columns
        # Minimum threshold to consider a withdrawal as an exit (8 ETH in gwei - minimum LEB bond)
        exit_threshold_gwei = 8 * 10**9

        # Separate exits from regular withdrawals
        # Exit must have is_exit=True AND be above the threshold (to filter out small skimmed rewards)
        principal_cap_gwei = 32 * 10**9

        if has_exit_flag:
            exit_mask = (
                (df_calc['record_type'] == 'withdrawal') &
                (df_calc['is_exit'] == True) &
                (df_calc['amount'] >= exit_threshold_gwei)
            )
            exits_df = df_calc[exit_mask].copy()
            withdrawals_df = df_calc[
                (df_calc['record_type'] == 'withdrawal') & ~exit_mask
            ].copy()

            # Handle excess above 32 ETH - split into principal (exit) and excess (withdrawal)
            # This matches invoice.py behavior
            excess_records = []
            for idx, row in exits_df.iterrows():
                if row['amount'] > principal_cap_gwei:
                    excess = row['amount'] - principal_cap_gwei
                    # Add excess as a withdrawal record
                    excess_record = row.copy()
                    excess_record['amount'] = excess
                    excess_records.append(excess_record)
                    # Cap the exit at 32 ETH
                    exits_df.at[idx, 'amount'] = principal_cap_gwei

            if excess_records:
                excess_df = pd.DataFrame(excess_records)
                withdrawals_df = pd.concat([withdrawals_df, excess_df], ignore_index=True)
        else:
            # Fallback for old data without is_exit flag
            exits_df = df_calc[(df_calc['record_type'] == 'withdrawal') & (df_calc['amount'] > principal_cap_gwei)].copy()
            withdrawals_df = df_calc[(df_calc['record_type'] == 'withdrawal') & (df_calc['amount'] <= principal_cap_gwei)].copy()

        proposals_df = df_calc[df_calc['record_type'] == 'proposal'].copy()

        # Apply reward adjustments for regular withdrawals and proposals
        if not withdrawals_df.empty:
            withdrawals_df['amount_adjusted'] = withdrawals_df.apply(
                lambda row: adjust_reward(row['amount'], row['validator_type']), axis=1
            )
            withdrawals_df['amount_eth'] = withdrawals_df['amount_adjusted'] / 10**9
            total_withdrawals = withdrawals_df['amount_eth'].sum()
        else:
            total_withdrawals = 0

        if not proposals_df.empty:
            proposals_df['amount_adjusted'] = proposals_df.apply(
                lambda row: adjust_reward(row['amount'], row['validator_type']), axis=1
            )
            proposals_df['amount_eth'] = proposals_df['amount_adjusted'] / 10**18
            total_proposals = proposals_df['amount_eth'].sum()
        else:
            total_proposals = 0

        # Calculate bonded principal for exits (not earnings)
        if not exits_df.empty:
            exits_df['amount_adjusted'] = exits_df.apply(
                lambda row: get_bonded_principal(row['amount'], row['validator_type']), axis=1
            )
            exits_df['amount_eth'] = exits_df['amount_adjusted'] / 10**9
            total_exits = exits_df['amount_eth'].sum()
            exit_count = len(exits_df)
        else:
            total_exits = 0
            exit_count = 0

        # Grand total excludes exits (exits are principal returns, not earnings)
        grand_total = total_withdrawals + total_proposals

        # Count validators
        total_validators = df_calc['validator_index'].nunique()

        # Rebuild df_calc for node breakdown (excluding exits from earnings breakdown)
        earnings_df = pd.concat([withdrawals_df, proposals_df]) if not withdrawals_df.empty or not proposals_df.empty else pd.DataFrame()

        if not earnings_df.empty:
            node_breakdown = earnings_df.groupby(['node', 'record_type']).agg({
                'amount_eth': 'sum',
                'validator_index': 'nunique'
            }).reset_index()
        else:
            node_breakdown = pd.DataFrame(columns=['node', 'record_type', 'amount_eth', 'validator_index'])

        return {
            'total_withdrawals': total_withdrawals,
            'total_proposals': total_proposals,
            'total_exits': total_exits,
            'exit_count': exit_count,
            'grand_total': grand_total,
            'total_validators': total_validators,
            'node_breakdown': node_breakdown,
            'record_count': len(df_calc) - exit_count  # Exclude exits from record count
        }

    def calculate_rate_of_return(self, df, total_earnings, epoch_duration):
        """Calculate rate of return based on validator type investment amounts."""
        # Calculate investment based on validator types
        validator_investments = df.groupby(['validator_index', 'validator_type']).first().reset_index()

        total_investment = 0
        for _, validator in validator_investments.iterrows():
            validator_type = validator['validator_type']
            try:
                validator_type = int(float(validator_type))
                # LEB8 validators (8-14): 8 ETH investment
                if 8 <= validator_type < 15:
                    total_investment += 8
                # LEB16 validators (16): 16 ETH investment
                elif 16 <= validator_type < 17:
                    total_investment += 16
                # Standard validators: 32 ETH investment
                else:
                    total_investment += 32
            except (ValueError, TypeError):
                # Default to 32 ETH if type cannot be parsed
                total_investment += 32

        # Calculate rate of return
        if total_investment > 0:
            rate_of_return = (total_earnings / total_investment) * 100

            # Annualized rate (assuming epoch_duration is in days)
            if epoch_duration > 0:
                annualized_rate = (rate_of_return * 365) / epoch_duration
            else:
                annualized_rate = 0
        else:
            rate_of_return = 0
            annualized_rate = 0

        return {
            'total_investment': total_investment,
            'rate_of_return': rate_of_return,
            'annualized_rate': annualized_rate
        }

    def epoch_to_date(self, epoch):
        """Convert epoch to approximate date (Genesis: 2020-12-01)."""
        genesis_timestamp = 1606824000  # Dec 1, 2020 12:00:00 UTC
        epoch_duration_seconds = 384  # ~6.4 minutes per epoch

        timestamp = genesis_timestamp + (epoch * epoch_duration_seconds)
        return datetime.fromtimestamp(timestamp)

    def create_professional_invoice(self, output_file, start_epoch, end_epoch,
                                  client_name="Valued Client", invoice_number=None):
        """Create professionally formatted Excel invoice."""

        # Load and calculate data
        df = self.load_data(start_epoch, end_epoch)
        earnings = self.calculate_earnings(df)

        # Calculate epoch duration in days
        start_date = self.epoch_to_date(start_epoch)
        end_date = self.epoch_to_date(end_epoch)
        epoch_duration = (end_date - start_date).days

        # Calculate ROI
        roi = self.calculate_rate_of_return(
            df,
            earnings['grand_total'],
            epoch_duration
        )

        # Generate invoice number if not provided
        if not invoice_number:
            invoice_number = f"INV-{datetime.now().strftime('%Y%m%d')}-{start_epoch}"

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoice"

        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15

        # Define styles
        header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        title_font = Font(name='Arial', size=18, bold=True)
        subtitle_font = Font(name='Arial', size=12, bold=True)
        normal_font = Font(name='Arial', size=10)

        header_fill = PatternFill(start_color=self.colors['header'], fill_type='solid')
        accent_fill = PatternFill(start_color=self.colors['accent'], fill_type='solid')

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        current_row = 1

        # Add logo if available
        if self.download_logo() and self.logo_path:
            try:
                img = image.Image(self.logo_path)
                img.width = 200
                img.height = 100
                ws.add_image(img, 'A1')
                current_row = 8
            except Exception as e:
                print(f"Could not add logo: {e}")

        # Company header
        ws[f'D{current_row}'] = "LI Blockchain"
        ws[f'D{current_row}'].font = title_font
        ws[f'D{current_row}'].alignment = Alignment(horizontal='right')

        current_row += 1
        ws[f'D{current_row}'] = "Validator Rewards Report"
        ws[f'D{current_row}'].font = subtitle_font
        ws[f'D{current_row}'].alignment = Alignment(horizontal='right')

        current_row += 3

        # Invoice details
        ws[f'A{current_row}'] = "INVOICE"
        ws[f'A{current_row}'].font = Font(name='Arial', size=16, bold=True)

        ws[f'D{current_row}'] = f"Invoice #: {invoice_number}"
        ws[f'D{current_row}'].font = subtitle_font
        ws[f'D{current_row}'].alignment = Alignment(horizontal='right')

        current_row += 2

        # Client and date info
        ws[f'A{current_row}'] = "Bill To:"
        ws[f'A{current_row}'].font = subtitle_font
        current_row += 1

        ws[f'A{current_row}'] = client_name
        ws[f'A{current_row}'].font = normal_font

        ws[f'D{current_row}'] = f"Date: {datetime.now().strftime('%B %d, %Y')}"
        ws[f'D{current_row}'].font = normal_font
        ws[f'D{current_row}'].alignment = Alignment(horizontal='right')

        current_row += 1
        ws[f'D{current_row}'] = f"Period: Epochs {start_epoch:,} - {end_epoch:,}"
        ws[f'D{current_row}'].font = normal_font
        ws[f'D{current_row}'].alignment = Alignment(horizontal='right')

        current_row += 1
        ws[f'D{current_row}'] = f"Duration: {epoch_duration} days"
        ws[f'D{current_row}'].font = normal_font
        ws[f'D{current_row}'].alignment = Alignment(horizontal='right')

        current_row += 3

        # Performance summary table
        ws[f'A{current_row}'] = "PERFORMANCE SUMMARY"
        ws[f'A{current_row}'].font = subtitle_font
        current_row += 1

        # Headers
        headers = ['Metric', 'Value', '', 'Investment Analysis', 'Value']
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=i, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        current_row += 1

        # Performance data
        performance_data = [
            ('Total Validators', f"{earnings['total_validators']:,}", '', 'Total Investment', f"{roi['total_investment']:.2f} ETH"),
            ('Total Records', f"{earnings['record_count']:,}", '', 'Total Earnings', f"{earnings['grand_total']:.6f} ETH"),
            ('Withdrawals', f"{earnings['total_withdrawals']:.6f} ETH", '', 'Rate of Return', f"{roi['rate_of_return']:.4f}%"),
            ('Proposals', f"{earnings['total_proposals']:.6f} ETH", '', 'Annualized Rate', f"{roi['annualized_rate']:.4f}%"),
        ]

        # Add exits row if there are any
        if earnings.get('total_exits', 0) > 0:
            performance_data.append(
                ('', '', '', 'Principal Returned (exits)', f"{earnings['total_exits']:.6f} ETH")
            )

        for row_data in performance_data:
            for i, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=i, value=value)
                cell.font = normal_font
                cell.border = border
                if i == 2 or i == 5:  # Value columns
                    cell.alignment = Alignment(horizontal='right')
            current_row += 1

        current_row += 2

        # Node breakdown if multiple nodes
        if len(earnings['node_breakdown']) > 1:
            ws[f'A{current_row}'] = "NODE BREAKDOWN"
            ws[f'A{current_row}'].font = subtitle_font
            current_row += 1

            # Node breakdown headers
            node_headers = ['Node', 'Type', 'Amount (ETH)', 'Validators']
            for i, header in enumerate(node_headers, 1):
                cell = ws.cell(row=current_row, column=i, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border

            current_row += 1

            # Node breakdown data
            for _, row in earnings['node_breakdown'].iterrows():
                node_data = [
                    row['node'],
                    row['record_type'].title(),
                    f"{row['amount_eth']:.6f}",
                    row['validator_index']
                ]

                for i, value in enumerate(node_data, 1):
                    cell = ws.cell(row=current_row, column=i, value=value)
                    cell.font = normal_font
                    cell.border = border
                    if i == 3 or i == 4:  # Numeric columns
                        cell.alignment = Alignment(horizontal='right')

                current_row += 1

        current_row += 3

        # Footer
        ws[f'A{current_row}'] = "Thank you for your business!"
        ws[f'A{current_row}'].font = Font(name='Arial', size=12, italic=True)

        # Create Details worksheet
        ws_details = wb.create_sheet(title="Details")

        # Set column widths for details
        ws_details.column_dimensions['A'].width = 12
        ws_details.column_dimensions['B'].width = 18
        ws_details.column_dimensions['C'].width = 12
        ws_details.column_dimensions['D'].width = 15
        ws_details.column_dimensions['E'].width = 15
        ws_details.column_dimensions['F'].width = 12
        ws_details.column_dimensions['G'].width = 20

        # Add headers for details
        detail_headers = ['Epoch', 'Validator Index', 'Node', 'Type', 'Amount (ETH)', 'Validator Type', 'Date/Time']
        for i, header in enumerate(detail_headers, 1):
            cell = ws_details.cell(row=1, column=i, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Add detail rows
        df_sorted = df.sort_values(['epoch', 'record_type', 'validator_index'])
        has_exit_flag = 'is_exit' in df.columns
        # Minimum threshold to consider a withdrawal as an exit (8 ETH in gwei - minimum LEB bond)
        exit_threshold_gwei = 8 * 10**9

        for idx, row in enumerate(df_sorted.itertuples(), 2):
            # Check if this is an exit (must have is_exit flag AND be above threshold)
            is_exit = False
            if row.record_type == 'withdrawal' and row.amount >= exit_threshold_gwei:
                if has_exit_flag and hasattr(row, 'is_exit'):
                    is_exit = row.is_exit == True
                elif row.amount > 32 * 10**9:
                    # Fallback for old data without is_exit flag
                    is_exit = True

            # Determine record type label and amount calculation
            if is_exit:
                # Exit: show bonded principal only, label as "Exit"
                record_type_label = "Exit"
                amount_adjusted = get_bonded_principal(row.amount, row.validator_type if hasattr(row, 'validator_type') else None)
                amount_eth = amount_adjusted / 10**9
            elif row.record_type == 'withdrawal':
                # Regular withdrawal: apply LEB adjustment
                record_type_label = "Withdrawal"
                amount_adjusted = adjust_reward(row.amount, row.validator_type if hasattr(row, 'validator_type') else None)
                amount_eth = amount_adjusted / 10**9
            else:
                # Proposal: apply LEB adjustment
                record_type_label = "Proposal"
                amount_adjusted = adjust_reward(row.amount, row.validator_type if hasattr(row, 'validator_type') else None)
                amount_eth = amount_adjusted / 10**18

            # Get datetime if available and format it
            datetime_str = ''
            if hasattr(row, 'datetime') and row.datetime:
                try:
                    # Handle different datetime formats
                    if isinstance(row.datetime, (int, float)):
                        # Unix timestamp - convert to datetime
                        dt = pd.to_datetime(row.datetime, unit='s')
                        datetime_str = dt.strftime('%m/%d/%Y %H:%M:%S')
                    elif isinstance(row.datetime, str):
                        dt = pd.to_datetime(row.datetime)
                        datetime_str = dt.strftime('%m/%d/%Y %H:%M:%S')
                    elif hasattr(row.datetime, 'strftime'):
                        datetime_str = row.datetime.strftime('%m/%d/%Y %H:%M:%S')
                    else:
                        datetime_str = str(row.datetime)
                except Exception as e:
                    datetime_str = str(row.datetime)

            # Convert validator type to ETH amount (8, 16, or 32)
            validator_type_eth = get_validator_type_label(row.validator_type if hasattr(row, 'validator_type') else None)

            detail_data = [
                row.epoch,
                row.validator_index,
                row.node,
                record_type_label,
                f"{amount_eth:.6f}",
                validator_type_eth,
                datetime_str
            ]

            for col, value in enumerate(detail_data, 1):
                cell = ws_details.cell(row=idx, column=col, value=value)
                cell.font = normal_font
                cell.border = border
                if col in [1, 2, 5]:  # Numeric columns
                    cell.alignment = Alignment(horizontal='right')

        # Save workbook
        wb.save(output_file)

        print(f"✅ Professional invoice generated: {output_file}")
        print(f"   📊 Total Earnings: {earnings['grand_total']:.6f} ETH")
        if earnings.get('total_exits', 0) > 0:
            print(f"   💰 Principal Returned (exits): {earnings['total_exits']:.6f} ETH")
        print(f"   📈 Rate of Return: {roi['rate_of_return']:.4f}%")
        print(f"   📅 Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")


def main():
    parser = argparse.ArgumentParser(description='Generate professional validator rewards invoice')
    parser.add_argument('start_epoch', type=int, help='Start epoch')
    parser.add_argument('end_epoch', type=int, help='End epoch')
    parser.add_argument('--parquet', default='rewards_data/rewards_master.parquet',
                       help='Input parquet file path')
    parser.add_argument('--output', help='Output Excel file path')
    parser.add_argument('--client', default='Valued Client', help='Client name')
    parser.add_argument('--invoice-number', help='Invoice number')

    args = parser.parse_args()

    # Generate output filename if not provided
    if not args.output:
        args.output = f"invoices/invoice_epochs_{args.start_epoch}_{args.end_epoch}.xlsx"

    try:
        generator = InvoiceGenerator(args.parquet)
        generator.create_professional_invoice(
            args.output,
            args.start_epoch,
            args.end_epoch,
            args.client,
            args.invoice_number
        )
    except Exception as e:
        print(f"Error generating invoice: {e}")


if __name__ == "__main__":
    main()